import glob
import os
import sys
from pathlib import Path
from typing import Generator

import openpyxl
import openpyxl.utils


class SearchResult:
    def __init__(self, file, sheet, row, col, val, complete_val=None) -> None:
        self.__file = file
        self.__sheet = sheet
        self.__row = row
        self.__col = openpyxl.utils.get_column_letter(col)
        self.__val = val
        self.__complete_val = complete_val if complete_val is not None else val

    def __str__(self):
        return f"File: {self.file}, Sheet: {self.sheet}, Cell: {self.cell}, Key: {self.value}, Is Partial Match: {self.is_partial_match}, Found Value: {self.found_value}"

    @property
    def file(self):
        return self.__file

    @property
    def sheet(self):
        return self.__sheet

    @property
    def cell(self):
        return self.__col + str(self.__row)

    @property
    def value(self):
        return self.__val

    @property
    def is_partial_match(self):
        return self.__complete_val != self.__val

    @property
    def found_value(self):
        return self.__complete_val


class DirectoryExcel:

    def __init__(
        self, root_directory: str = str(Path.home()), recursive: bool = False
    ) -> None:
        self.__root_dir = os.path.abspath(root_directory)
        self.__recursive = recursive
        self.excel_files = []
        self.find_excel_files()

    @property
    def root_dir(self):
        return self.__root_dir

    @root_dir.setter
    def root_dir(self, new_root_dir):
        self.excel_files = []
        self.__root_dir = os.path.abspath(new_root_dir)
        self.find_excel_files()

    @property
    def recursive(self):
        return self.__recursive

    @recursive.setter
    def recursive(self, new_recursive_policy):
        self.excel_files = []
        self.__recursive = new_recursive_policy
        self.find_excel_files()

    def find_excel_files(self):
        extensions = [
            "*.xlsx",
            "*.xls",
            "*.xlsm",
            "*.xlsb",
            "*.xltx",
            "*.xltm",
            "*.xlt",
            "*.xml",
            "*.ods",
        ]
        for extension in extensions:
            files_found = []
            if self.__recursive:
                search_pattern = os.path.join(self.__root_dir, "**", extension)
                files_found = glob.glob(search_pattern, recursive=True)
            else:
                search_pattern = os.path.join(self.__root_dir, extension)
                files_found = glob.glob(search_pattern, recursive=False)
            self.excel_files.extend(
                [
                    file
                    for file in files_found
                    if not os.path.basename(file).startswith("~$")
                ]
            )

    def search_keyword(
        self, keyword: str, exact_match=False
    ) -> Generator[SearchResult, None, None]:
        if keyword == "":
            return
        for excel_file in self.excel_files:
            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                for ws in wb.worksheets:
                    for col in ws.iter_cols():
                        for cell in col:
                            cell_value = str(cell.value)
                            if keyword == cell_value:
                                yield SearchResult(
                                    excel_file, ws.title, cell.row, cell.column, keyword
                                )
                            elif not exact_match and keyword in cell_value:
                                yield SearchResult(
                                    excel_file,
                                    ws.title,
                                    cell.row,
                                    cell.column,
                                    keyword,
                                    cell_value,
                                )
            except Exception as e:
                print(f"Error opening {excel_file}: {e}")


if __name__ == "__main__":
    recursive = False
    if len(sys.argv) > 1:
        recursive = sys.argv[1].lower() == "true"
    dir_excel = DirectoryExcel(".", recursive)
    print(dir_excel.root_dir)
    for excel_file in dir_excel.excel_files:
        print(excel_file)
